import os
import pandas as pd
from openpyxl.utils import get_column_letter
import json
import math

# Input
filename = os.environ.get("INPUT_EXCEL_PATH")
if not filename:
    raise RuntimeError("INPUT_EXCEL_PATH is required for pdf9.py")

filepath = filename if os.path.isabs(filename) else os.path.join(os.getcwd(), filename)

try:
    df = pd.read_excel(filepath, skiprows=4)
    df['Betrag'] = df['Betrag'].astype(str).str.replace(',', '.').astype(float)
    df['Erledigt'] = pd.to_datetime(df['Erledigt'], format='%d.%m.%Y', errors='coerce')
    print("Datei geladen:", filepath)
except FileNotFoundError:
    raise RuntimeError(f"Datei nicht gefunden: {filepath}")

new_df = df[['Hersteller', 'Fahrgestellnummer', 'Fahrzeugtyp', 'Erledigt', 'Betrag']].copy()
new_df.rename(columns={
    'Fahrgestellnummer': 'VIN',
    'Fahrzeugtyp':       'Fahrzeug'
}, inplace=True)

new_df = new_df[new_df['Fahrzeug'].notna()]
new_df = new_df[new_df['Fahrzeug'].str.strip() != ""]

# Load DB
ca3_url = os.getenv("CA3_URL_Excel")
rrm_url = os.getenv("RRM_URL_Excel")

if not ca3_url or not rrm_url:
    config_path = os.path.join(os.getcwd(), "config.json")
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        ca3_url = ca3_url or config.get("CA3_URL_Excel", "")
        rrm_url = rrm_url or config.get("RRM_URL_Excel", "")

df_ca3 = pd.read_json(ca3_url)
df_rrm = pd.read_json(rrm_url)

df_ca3["Auftraggeber"] = "CA3"
df_rrm["Auftraggeber"] = "RRM"
df_metabase = pd.concat([df_ca3, df_rrm], ignore_index=True)

print("DB geladen:", len(df_metabase), "Einträge")

# Helpers
def is_empty(v):
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return False

# Duplikate
dup_vins = set(new_df[new_df.duplicated(subset=["VIN"], keep=False)]["VIN"])

# Merge by VIN (keep highest invoice per VIN)
df_metabase_dedup = (
    df_metabase
    .sort_values("invoice", ascending=False)
    .drop_duplicates(subset=["vin"], keep="first")
)

merged_df = pd.merge(
    new_df,
    df_metabase_dedup,
    left_on="VIN",
    right_on="vin",
    how="left"
)

# Vergleich
merged_df["VIN_vergleich"] = merged_df["vin"].apply(
    lambda v: "NOK" if is_empty(v) else "OK"
)

merged_df["EÜbernahme_vergleich"] = merged_df.apply(
    lambda row: (
        "OK" if row["VIN_vergleich"] == "OK" and not is_empty(row.get("EÜbernahme"))
        else "NOK"
    ),
    axis=1
)

# Bemerkungen
def bemerkungen_logic(row):
    parts = []
    if row["VIN"] in dup_vins:
        parts.append("Dublikat, prüfen")
    if row["VIN_vergleich"] == "NOK":
        parts.append("Transportauftrag nicht gefunden. Bitte WFPs 2010, 9010, 9020 manuell prüfen")
    elif row["EÜbernahme_vergleich"] == "NOK":
        parts.append("WFP EÜbernahme nicht aktiviert, bitte prüfen")
    else:
        parts.append("OK, wurde weiterverrechnet")
    return " | ".join(parts)

merged_df["Bemerkungen"] = merged_df.apply(bemerkungen_logic, axis=1)

# Output: ALL rows
df_final = merged_df[[
    "Hersteller", "VIN", "Fahrzeug", "Erledigt", "Betrag",
    "Auftraggeber", "EÜbernahme", "Bemerkungen"
]].rename(columns={"EÜbernahme": "WFP_EÜbernahme_Preis"}).copy()

BASE_DIR = os.environ.get("BASE_DIR")
if not BASE_DIR:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

final_file = os.path.join(BASE_DIR, "Fehlerreport.xlsx")

with pd.ExcelWriter(final_file, engine="openpyxl") as writer:
    df_final.to_excel(writer, index=False, sheet_name="Fehlerreport")
    ws = writer.sheets["Fehlerreport"]
    for i in range(1, len(df_final.columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 25
    ws.column_dimensions[get_column_letter(len(df_final.columns))].width = 50  # Bemerkungen

print(f"Fehlerreport gespeichert: {final_file}")
print(f"Fehler gesamt: {len(df_final)}")