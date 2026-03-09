import os
import pandas as pd
from openpyxl.utils import get_column_letter
import json
import math

# 1. Load input file
filename = os.environ.get("INPUT_EXCEL_PATH")
if not filename:
    raise RuntimeError("INPUT_EXCEL_PATH is required for pdf10.py")

filepath = filename if os.path.isabs(filename) else os.path.join(os.getcwd(), filename)

try:
    # Read Excel, skipping the first 4 rows of the header
    df = pd.read_excel(filepath, skiprows=4)
    # Perform standard data cleaning on the input DataFrame
    df['Betrag'] = df['Betrag'].astype(str).str.replace(',', '.').astype(float)
    df['Eingang'] = pd.to_datetime(df['Eingang'], format='%d.%m.%Y', errors='coerce')
    df['Ausgang'] = pd.to_datetime(df['Ausgang'], format='%d.%m.%Y', errors='coerce')
    
    # Keep Order-Nr as a clean string for reporting
    df['Order-Nr.'] = df['Order-Nr.'].astype(str).str.strip()
    
    print(f"File loaded: {filepath}")
except FileNotFoundError:
    raise RuntimeError(f"File not found: {filepath}")

# Prepare the input data subset for processing
new_df = df[['Hersteller', 'Fahrgestellnummer', 'Order-Nr.', 'Fahrzeugtyp', 'Eingang', 'Ausgang', 'Betrag']].copy()
new_df.rename(columns={
    'Fahrgestellnummer': 'VIN',
    'Fahrzeugtyp':       'Fahrzeug',
    'Order-Nr.':         'OrderNr'
}, inplace=True)

new_df = new_df[new_df['Fahrzeug'].notna()]
new_df = new_df[new_df['Fahrzeug'].str.strip() != ""]

# VIN NORMALIZATION for lookup (applied to input data)
# This converts all VINs to uppercase and removes any spaces to prevent mismatching
new_df['VIN'] = new_df['VIN'].astype(str).str.strip().str.upper()

# 2. Load Database (Metabase) from environment variables or config.json
ca3_url = os.getenv("CA3_Lagerhandling")
rrm_url = os.getenv("RRM_Lagerhandling")

# Try to load credentials from config.json if environment variables are empty
if not ca3_url or not rrm_url:
    config_path = os.path.join(os.getcwd(), "config.json")
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        ca3_url = ca3_url or config.get("CA3_Lagerhandling", "")
        rrm_url = rrm_url or config.get("RRM_Lagerhandling", "")

# Initial empty DataFrame for the database
df_metabase = pd.DataFrame()

# Attempt to fetch and load data from the DB URLs
try:
    if not ca3_url or not rrm_url:
        print("Warning: CA3_Lagerhandling or RRM_Lagerhandling URL is missing.")
    else:
        df_ca3 = pd.read_json(ca3_url)
        df_rrm = pd.read_json(rrm_url)
        
        # Normalize DB column names to lowercase for consistency
        df_ca3.columns = [col.lower() for col in df_ca3.columns]
        df_rrm.columns = [col.lower() for col in df_rrm.columns]

        # Add the Auftraggeber column to identify the source
        df_ca3["Auftraggeber"] = "CA3"
        df_rrm["Auftraggeber"] = "RRM"
        # Concatenate both data sources into a single database DataFrame
        df_metabase = pd.concat([df_ca3, df_rrm], ignore_index=True)
    
    if not df_metabase.empty:
        # VIN NORMALIZATION for lookup (applied to database)
        # Ensure database VINs match the format used for input data normalization
        df_metabase['vin'] = df_metabase['vin'].astype(str).str.strip().str.upper()
        print(f"DB loaded: {len(df_metabase)} entries total")
    else:
        print("Warning: Database is empty. All VIN lookups will fail.")

except Exception as e:
    print(f"Error loading database: {e}")
    # In case of error, continue with an empty DataFrame containing required columns
    df_metabase = pd.DataFrame(columns=['vin', 'invoice', 'Auftraggeber', 'wfp4500'])

# 3. Refactored Vectorized Lookup (Merged lookup strategy instead of looping .apply())
# This approach replaces the buggy `apply` logic.

# Create a lookup table by keeping unique VINs and prioritizing the best entry per VIN
db_lookup = pd.DataFrame(columns=['vin', 'invoice', 'Auftraggeber', 'wfp4500'])

if not df_metabase.empty:
    db_lookup = df_metabase.copy()
    
    # Safely convert 'invoice' to numeric format for sorting
    # This ensures that 'sort_values' prioritizes numerically higher invoices correctly.
    def robust_num_convert(v):
        if pd.notna(v) and v != '':
            try:
                # Assuming 'invoice' can be converted to an integer, e.g., '12345'
                return int(float(v))
            except (ValueError, OverflowError):
                # For non-numeric invoices, treat as very low priority
                return 0
        return 0

    # Sort and filter df_metabase to have one UNIQUE entry per VIN
    # Sort by VIN ascending, then by numeric invoice value ascending
    # 'drop_duplicates' keeping 'last' will preserve the entry with the highest invoice per VIN.
    if 'invoice' in db_lookup.columns:
        db_lookup['_sort_invoice'] = db_lookup['invoice'].apply(robust_num_convert)
        db_lookup = db_lookup.sort_values(['vin', '_sort_invoice'], ascending=[True, True])
        db_lookup = db_lookup.drop_duplicates(subset='vin', keep='last')
        db_lookup = db_lookup.drop(columns='_sort_invoice') # remove helper column
    else:
        db_lookup = db_lookup.drop_duplicates(subset='vin', keep='first')

# PERFORM THE VECTORIZED LOOKUP (Left Merge)
# Merge input data (new_df) with prepared database lookup table (db_lookup)
# The merge joins where new_df['VIN'] matches db_lookup['vin']
# 'left' merge ensures all rows from new_df are kept; missing lookups will be filled with NaN.
merged_df = pd.merge(
    new_df.reset_index(drop=True),
    db_lookup.reset_index(drop=True),
    left_on='VIN',
    right_on='vin',
    how='left'
)

# 4. Validation and Reporting Logic
# The lookup results from `wfp4500` will naturally contain `NaN` where matches were not found.
merged_df["VIN_vergleich"] = merged_df["vin"].apply(
    lambda v: "NOK" if (v is None or pd.isna(v) or str(v).strip() == "" or str(v) == 'nan') else "OK"
)

merged_df["WFP4500_vergleich"] = merged_df.apply(
    lambda row: "OK" if row["VIN_vergleich"] == "OK" and pd.notna(row.get("wfp4500")) else "NOK",
    axis=1
)

# Duplicate VIN detection within the input file itself
dup_vins = set(new_df[new_df.duplicated(subset=["VIN"], keep=False)]["VIN"])

def bemerkungen_logic(row):
    parts = []
    if row["VIN"] in dup_vins:
        parts.append("Duplicate in file")
    
    # Report generation using merged and validated data
    if row["VIN_vergleich"] == "NOK":
        parts.append("Transport order not found. Please check manually")
    else:
        if row["WFP4500_vergleich"] == "NOK":
            parts.append("WFP 4500 not activated")
        else:
            parts.append("OK, processed")
    return " | ".join(parts)

# Apply reporting logic vectorized across the merged DataFrame
merged_df["Bemerkungen"] = merged_df.apply(bemerkungen_logic, axis=1)

# 5. Final Report Generation (ALL Rows)
# Select only the relevant columns and rename them to the expected output format
df_final = merged_df[[
    "Hersteller", "VIN", "OrderNr", "Fahrzeug", "Eingang", "Ausgang", "Betrag",
    "Auftraggeber", "wfp4500", "Bemerkungen"
]].rename(columns={
    "OrderNr": "Order-Nr",
    "wfp4500": "WFP_4500_Preis"
}).copy()

# Base directory handling from environment variable or current work directory
BASE_DIR = os.environ.get("BASE_DIR", os.getcwd())
final_file = os.path.join(BASE_DIR, "Fehlerreport.xlsx")

# Save the final DataFrame to an Excel file with adjusted column widths
with pd.ExcelWriter(final_file, engine="openpyxl") as writer:
    df_final.to_excel(writer, index=False, sheet_name="Fehlerreport")
    ws = writer.sheets["Fehlerreport"]
    # Dynamic column widths for readability in Excel
    for i in range(1, len(df_final.columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 25
    # Extra width for the Bemerkungen column
    ws.column_dimensions[get_column_letter(len(df_final.columns))].width = 60

print(f"Success! Final report saved to: {final_file}")
print(f"Total entries processed: {len(df_final)}")