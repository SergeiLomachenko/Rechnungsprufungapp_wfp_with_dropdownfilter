import re
import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os 
import json
import numpy as np

# PDF-path
excel_path = "invoice.xlsx"

df = pd.read_excel("invoice.xlsx", skiprows=5) 
df_5 = df.copy()

df_temp = df.replace(r'^\s*$', pd.NA, regex=True)
empty_rows = df_temp.isna().all(axis=1) 
if empty_rows.any():
    first_empty_index = empty_rows.idxmax()
    df = df.loc[:first_empty_index - 1]
else:
    pass
df = df.reset_index(drop=True)
df = df.dropna(how='all')

# changes to PLZ
def clean_plz(series):
    numeric = pd.to_numeric(series, errors='coerce').fillna(0)
    # making to int, then to string
    return numeric.astype(int).astype(str)
df["Absender PLZ"] = clean_plz(df["Absender PLZ"])
df["Empfaenger PLZ"] = clean_plz(df["Empfaenger PLZ"])

for col in ["Ansatz", "Faktor", "Betrag"]:
    df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '.'), errors='coerce')

invoice_nr = (
    df['Ordernummer']
    .fillna('')
    .astype(str)
    .str.extract(r'^\s*([^\s/]+)')  # extract all till the first backspase
    .squeeze()
    .str.strip()
)
lengths = invoice_nr.str.len()

# Auftraggeber kalkulieren
auftraggeber = lengths.map({
    5: "RRM",
    6: "CA3"
}).fillna("Fehler")

new_df = pd.DataFrame({
    "InvoiceNr": df["Ordernummer"],
    "Invoiceshort" : invoice_nr,
    'Auftraggeber': auftraggeber,
    'VIN': df['Fahrgestellnummer'], 
    'Model': df['Fahrzeugtyp'], 
    "Ansatz": df["Ansatz"],
    "Faktor": df["Faktor"],
    "Nebenkosten": df["Nebenkosten"],
    "Total": df["Betrag"],
    "Loadingcity": df["Absenderort"],
    "Delivercity": df["Empfaengerort"],
    "Absender" : df["Absender"],
    "Empfaenger": df["Empfaenger"],
    "Absender PLZ" : df["Absender PLZ"],
    "Empfaenger PLZ": df["Empfaenger PLZ"]
})

nebenkosten_df = new_df[new_df['Nebenkosten'].notna()]

# loading envs
ca3_daten = os.getenv("CA3_URL_Excel")
rrm_daten = os.getenv("RRM_URL_Excel")

if not ca3_daten:
    config_path = os.path.join(os.getcwd(), "config.json")
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        ca3_daten = config.get("CA3_URL_Excel", "")

if not rrm_daten:
    config_path = os.path.join(os.getcwd(), "config.json")
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        rrm_daten = config.get("RRM_URL_Excel", "")

db_ca3 = pd.read_json(ca3_daten)
db_rrm = pd.read_json(rrm_daten) 
df_ca3 = db_ca3.copy()
df_rrm = db_rrm.copy()


# Below are all the functions that were used to compare the invoice-data with metabase-database
def compare_text_values(v1, v2):
    """
    Vergleicht zwei Textwerte.
    Fehlt ein Wert (pd.isna), wird er als leerer String berücksichtigt.
    Zusätzlich werden unsichtbare Leerzeichen entfernt.
    Beide Werte werden getrimmt und in Kleinbuchstaben umgewandelt.
    """
    def clean(val):
        if pd.isna(val):
            return ""
        # Приводим к строке, убираем неразрывные пробелы и обычные пробелы по краям
        s = str(val).replace("\xa0", " ").strip()
        # Убираем возможные .0 у целых чисел (если были float)
        if s.endswith(".0"):
            s = s[:-2]
        return s.lower()

    s1 = clean(v1)
    s2 = clean(v2)
    return "OK" if s1 == s2 else "NOK"



# def compare_text_values(v1, v2):
#     """
#     Vergleicht zwei Textwerte.
#     Fehlt ein Wert (pd.isna), wird er als leerer String berücksichtigt.
#     Zusätzlich werden unsichtbare Leerzeichen entfernt.
#     Beide Werte werden getrimmt und in Kleinbuchstaben umgewandelt.
#     """
#     s1 = "" if pd.isna(v1) else str(v1).replace("\xa0", "").strip().lower()
#     s2 = "" if pd.isna(v2) else str(v2).replace("\xa0", "").strip().lower()
#     return "OK" if s1 == s2 else "NOK"

def longest_common_substring(s1, s2):
    m, n = len(s1), len(s2)
    dp = [[0]*(n+1) for _ in range(m+1)]
    longest = 0
    for i in range(1, m+1):
        for j in range(1, n+1):
            if s1[i-1] == s2[j-1]:
                dp[i][j] = dp[i-1][j-1] + 1
                if dp[i][j] > longest:
                    longest = dp[i][j]
            else:
                dp[i][j] = 0
    return longest

def compare_city(val1, val2):
    # Konvertiere die Eingabewerte in Kleinbuchstaben,
    # entferne unsichtbare Leerzeichen (z. B. non-breaking spaces) und trimme den Whitespace.
    s1 = "" if pd.isna(val1) else str(val1).replace("\xa0", "").strip().lower()
    s2 = "" if pd.isna(val2) else str(val2).replace("\xa0", "").strip().lower()
    
    # Falls beide Werte leer sind, werden sie als gleich angesehen.
    if s1 == "" and s2 == "":
        return "OK"
    
    # Sonderregel: Wenn ein Wert "nebikon" und der andere "altishofen" (oder umgekehrt) ist,
    # werden diese als gleich betrachtet.
    if (s1 == "nebikon" and s2 == "altishofen") or (s1 == "altishofen" and s2 == "nebikon"):
        return "OK"
    
    # Wenn die Länge der längsten gemeinsamen Teilzeichenkette mindestens 5 Zeichen beträgt,
    # werden die Werte als ähnlich betrachtet.
    if (
        longest_common_substring(s1, s2) >= 3 
        or (longest_common_substring(s1, s2) >= 2 and s1 == "au") 
        or (longest_common_substring(s1, s2) >= 2 and s2 == "au")
        ):
        return "OK"
    
    # Ansonsten gelten die Werte als unterschiedlich.
    return "NOK"
# End of all the functions that were used to compare the invoice-data with metabase-database

# Liste zur Speicherung der Vergleichsergebnisse erstellen
compare_results = []

for idx, row in new_df.iterrows():
    Invoiceshort = str(row["Invoiceshort"]).strip()
    
    # Auswahl der Referenzdatei anhand der Länge der InvoiceNr:
    # - 5-stellige InvoiceNr → rrm.xlsx
    # - 6-stellige InvoiceNr → ca3.xlsx
    if len(Invoiceshort) == 5:
        df_reference = df_rrm
    elif len(Invoiceshort) == 6:
        df_reference = df_ca3
    else:
        df_reference = None

    # Suche der passenden Zeile in der Referenzdatei anhand der Spalte "invoice"
    if df_reference is not None:
        matching_rows = df_reference[df_reference["invoice"].astype(str).str.strip() == Invoiceshort]
        if len(matching_rows) > 0:
            ref = matching_rows.iloc[0]
        else:
            ref = None
    else:
        ref = None

    # Falls eine Referenzzeile gefunden wurde, erfolgt der Vergleich.
    def compare_null_logic(val1, val2):
        # считаем пустыми: None, "", пробелы, NaN
        def is_empty(v):
            if v is None:
                return True
            if isinstance(v, str) and v.strip() == "":
                return True
            try:
                import math
                if isinstance(v, float) and math.isnan(v):
                    return True
            except:
                pass
            return False

        if is_empty(val1) and is_empty(val2):
            return "OK"
        elif not is_empty(val1) and not is_empty(val2):
            return "OK"
        else:
            return "NOK"
    

    if ref is not None:
        auftraggeber_vergleich = compare_text_values(row["Auftraggeber"], ref["Auftraggeber"]) 
        vin_vergleich = compare_text_values(row["VIN"], ref["vin"]) 
        loadingcity_vergleich = compare_city(row["Loadingcity"], ref["loadingcity"]) 
        delivercity_vergleich = compare_city(row["Delivercity"], ref["delivercity"])
        loadingzipcode_vergleich = compare_text_values(row["Absender PLZ"], ref["loadingzipcode"])
        deliverzipcode_vergleich = compare_text_values(row["Empfaenger PLZ"], ref["deliverzipcode"])
        faktor_vergleich = compare_null_logic(row["Faktor"], ref["Faktor"])
        # transportrpeis_vergleich = compare_null_logic(row["Total"], ref["Faktor"])

        # telavis_vergleich
        # def is_emptytel(val):
        #     return val is None or (isinstance(val, str) and val.strip() == "")
        # if is_emptytel(row["Terminverein. Absender CarAukt"]):
        #     telavis_vergleich = "OK"
        # else:
        #     telavis_vergleich = compare_null_logic(new_df["Terminverein. Absender CarAukt"], ref["Terminzuschlag"])

        # seilwinde_vergleich
        # def is_emptyseil(val):
        #     return val is None or (isinstance(val, str) and val.strip() == "")
        # if is_emptyseil(row["Seilwinde"]):
        #     seilwinde_vergleich = "OK"
        # else:
        #     seilwinde_vergleich = compare_null_logic(new_df["Seilwinde"], ref["Seilwinde"])
        
        # compare Seilwindeintransport
        # def is_emptyseiltr(val):
        #     return val is None or (isinstance(val, str) and val.strip() == "")
        # if is_emptyseiltr(new_df["Seilwinde"]):
        #     seilwinde_transport_vergleich = "OK"
        # else:
        #     seilwinde_transport_vergleich = compare_null_logic(new_df["Seilwinde"], ref["Seilwindeintransport"])

        # terminzuschlag_vergleich
        # def is_emptytermin(val):
        #     return val is None or (isinstance(val, str) and val.strip() == "")
        # if is_emptytermin(new_df["Terminzuschlag"]):
        #     terminzuschlag_vergleich = "OK"
        # else:
        #     terminzuschlag_vergleich = compare_null_logic(new_df["Terminzuschlag"], ref["Terminzuschlag"])

        # efahrzeug_vergleich
        # below is new function for uebernahme comparison
        # def is_empty(val):
        #     return val is None or (isinstance(val, str) and val.strip() == "")
        # if is_empty(new_df["Car Auktion Protokoll"]):
        #     efahrzeug_vergleich = "OK"
        # else:
        #     efahrzeug_vergleich = compare_null_logic(new_df["Car Auktion Protokoll"], ref["EÜbernahme"])

        # leerfahrt_vergleich
        # def is_emptyleer(val):
        #     return val is None or (isinstance(val, str) and val.strip() == "")
        # if is_empty(new_df["LEERFAHRT"]):
        #     leerfahrt_vergleich = "OK"
        # else:
        #     leerfahrt_vergleich = compare_null_logic(new_df["LEERFAHRT"], ref["Leerfahrt"])

    else:
        auftraggeber_vergleich = vin_vergleich = loadingcity_vergleich = delivercity_vergleich = "NOK" 
        loadingzipcode_vergleich = deliverzipcode_vergleich = "NOK"
        faktor_vergleich = transportrpeis_vergleich = "NOK"
        # telavis_vergleich = seilwinde_vergleich = seilwinde_transport_vergleich = terminzuschlag_vergleich = "NOK"
        # efahrzeug_vergleich = leerfahrt_vergleich = "NOK"
        


# Zusammenstellung der neuen Zeile mit den Originaldaten aus file2 und den Vergleichsergebnissen
    new_row = {
        "InvoiceNr": row["InvoiceNr"],
        "Invoiceshort" : row["Invoiceshort"],
        "Auftraggeber": row["Auftraggeber"],  # Auftraggeber aus file2
        "VIN": row["VIN"],
        "Model": row["Model"],
        "Faktor": row["Faktor"],
        "Total": row["Total"],
        "Loadingcity": row["Loadingcity"],
        "Delivercity": row["Delivercity"],
        "Absender PLZ": row["Absender PLZ"],
        "Empfaenger PLZ": row["Empfaenger PLZ"], 
        # Vergleichsergebnisse
        "AuftraggeberVergleich": auftraggeber_vergleich,
        "VINVergleich": vin_vergleich,
        "Absender PLZ Vergleich": loadingzipcode_vergleich,
        "Empfaenger PLZ Vergleich": deliverzipcode_vergleich,
        "LoadingcityVergleich": loadingcity_vergleich,
        "DelivercityVergleich": delivercity_vergleich,
        "TransportWFP_aktiviert_Vergleich": faktor_vergleich
        # "TransportrpeisVergleich": transportrpeis_vergleich

        # "TelavisVergleich": telavis_vergleich,
        # "SeilwindeVergleich": seilwinde_vergleich,
        # "SeilwindeTransportVergleich": seilwinde_transport_vergleich,
        # "TerminzuschlagVergleich": terminzuschlag_vergleich,
        # "E-FahrzeugVergleich": efahrzeug_vergleich,
        # "LeerfahrtVergleich": leerfahrt_vergleich,
        # "Bemerkungen": bemerkungen_text
    }
    compare_results.append(new_row)

# Erstellen eines neuen DataFrames für file4 mit den Vergleichsergebnissen
df_vergleich = pd.DataFrame(compare_results)
# columns we need
needed_columns = [
    "InvoiceNr",
    "Invoiceshort",
    "Auftraggeber",
    "VIN",
    "Model",
    "Faktor",
    "Total",
    "Loadingcity",
    "Delivercity",
    "Absender PLZ",
    "Empfaenger PLZ",
    "AuftraggeberVergleich",
    "VINVergleich",
    "Absender PLZ Vergleich", 
    "Empfaenger PLZ Vergleich",
    "LoadingcityVergleich",
    "DelivercityVergleich",
    "TransportWFP_aktiviert_Vergleich"
]

# Оставить только нужные колонки (и в нужном порядке)
df_vergleich = df_vergleich[needed_columns]

# deleting columns if exist 
cols_to_drop = ["FactorVergleich", "TransportrpeisVergleich"]
for col in cols_to_drop:
    if col in df_vergleich.columns:
        df_vergleich = df_vergleich.drop(columns=[col])


# making sure the type of the columns is okey
df_vergleich = df_vergleich.astype({
    "InvoiceNr": "string",
    "Invoiceshort": "string",
    "Auftraggeber": "string",
    "VIN": "string",
    "Model": "string",
    "Faktor": "float",
    "Total": "float",
    "Absender PLZ": "string",
    "Empfaenger PLZ": "string", 
    "Loadingcity": "string",
    "Delivercity": "string",
    "AuftraggeberVergleich": "string",
    "VINVergleich": "string",
    "Absender PLZ Vergleich": "string",
    "Empfaenger PLZ Vergleich": "string",
    "LoadingcityVergleich": "string",
    "DelivercityVergleich": "string",
    "TransportWFP_aktiviert_Vergleich": "string"
    # "TransportrpeisVergleich": "string"
})


# Conditions for step_1: Anzahl CA3, RRM, Fehler:
ca3_count = df_vergleich[
    (df_vergleich["Auftraggeber"] == "CA3") &
    (df_vergleich["Faktor"] != 0) &
    (df_vergleich["AuftraggeberVergleich"] == "OK")
].shape[0]

rrm_count = df_vergleich[
    (df_vergleich["Auftraggeber"] == "RRM") &
    (df_vergleich["Faktor"] != 0) &
    (df_vergleich["AuftraggeberVergleich"] == "OK")
].shape[0]

# fehler_count = df_vergleich[
#     (df_vergleich["Auftraggeber"] == "Fehler") &
#     (df_vergleich["Faktor"] != 0) &
#     (df_vergleich["AuftraggeberVergleich"] == "NOK")
# ].shape[0]

fehler_count = df_vergleich[
    ((df_vergleich["Auftraggeber"] == "Fehler") & (df_vergleich["Faktor"] != 0)) & (df_vergleich["AuftraggeberVergleich"] == "NOK") |
    ((df_vergleich["Auftraggeber"] == "CA3") & (df_vergleich["Faktor"] != 0) & (df_vergleich["AuftraggeberVergleich"] == "NOK")) |
    ((df_vergleich["Auftraggeber"] == "RRM") & (df_vergleich["Faktor"] != 0) & (df_vergleich["AuftraggeberVergleich"] == "NOK"))
].shape[0]


# creating new dataFrame for step_1
step1_df = pd.DataFrame({
    "CA3": [ca3_count],
    "RRM": [rrm_count],
    "Fehler": [fehler_count]
})

# cleen the df_vergleich to make less columns
columns_to_keep_step_1 = ["Invoiceshort", "VIN", "Model", "Faktor", "Total", "Loadingcity", "Delivercity", "Auftraggeber", "AuftraggeberVergleich"]
df_vergleich_cleen = df_vergleich[columns_to_keep_step_1]

# Conditions for step_1.1-1.3: CA3 RRM Fehler FZG
df_ca3 = df_vergleich_cleen[
    (df_vergleich_cleen["Auftraggeber"] == "CA3") &
    (df_vergleich_cleen["Faktor"] != 0) &
    (df_vergleich_cleen["AuftraggeberVergleich"] == "OK")
]

df_rrm = df_vergleich_cleen[
    (df_vergleich_cleen["Auftraggeber"] == "RRM") &
    (df_vergleich_cleen["Faktor"] != 0) &
    (df_vergleich_cleen["AuftraggeberVergleich"] == "OK")
]

df_fehler = df_vergleich_cleen[
    ((df_vergleich_cleen["Auftraggeber"] == "Fehler") & (df_vergleich_cleen["Faktor"] != 0)) & (df_vergleich_cleen["AuftraggeberVergleich"] == "NOK") |
    ((df_vergleich_cleen["Auftraggeber"] == "CA3") & (df_vergleich_cleen["Faktor"] != 0) & (df_vergleich_cleen["AuftraggeberVergleich"] == "NOK")) |
    ((df_vergleich_cleen["Auftraggeber"] == "RRM") & (df_vergleich_cleen["Faktor"] != 0) & (df_vergleich_cleen["AuftraggeberVergleich"] == "NOK"))
]

# df_fehler = df_vergleich_cleen[
#     (df_vergleich_cleen["Auftraggeber"] == "Fehler") &
#     (df_vergleich_cleen["Faktor"] != 0) 

# ]

# creating new dataFrame for step_1_1
step_1_1_df = df_fehler
# creating new dataFrame for step_1_2
step_1_2_df = df_ca3
# creating new dataFrame for step_1_3
step_1_3_df = df_rrm

# creating a dataframe for step_1_4 zusätzliche Info
df_zusatz = df_vergleich[
    (df_vergleich["Faktor"] != 0) &
    (df_vergleich["AuftraggeberVergleich"] == "OK") &
    (
    (df_vergleich["VINVergleich"] == "NOK") | 
    (df_vergleich["Absender PLZ Vergleich"] == "NOK") | 
    (df_vergleich["Empfaenger PLZ Vergleich"] == "NOK") | 
    (df_vergleich["LoadingcityVergleich"] == "NOK") | 
    (df_vergleich["DelivercityVergleich"] == "NOK") | 
    (df_vergleich["DelivercityVergleich"] == "NOK") |
    (df_vergleich["TransportWFP_aktiviert_Vergleich"] == "NOK")
    )
]

# STEP 2: Überprüfung von Beträgen, Suche nach nicht gerechtfertigten und falschen
def check_total_amounts(df):
    # list of okey Beträgen
    valid_totals = [126, 156.8, 189, 207.75, 235.20, 294.35, 
                    311.65, 311.6, 313.60, 378, 392, 415.5, 
                    127.25, 158.35, 190.90, 190.85, 209.85, 237.55, 237.5,
                    314.80, 314.75, 395.90, 395.85, 419.70, 297.30, 381.80, 
                    385.95, 390.70, 416.95]
                    #Zeile oben ist dür die Motorräder
    df['Total'] = pd.to_numeric(df['Total'], errors='coerce')
    df['Total_rounded'] = df['Total'].round(2)
    # round to 2 decimals and check if in the valid_totals
    # df['Betrag okey'] = df['Total'].round(2).isin([round(v, 2) for v in valid_totals]).map({True: 'OK', False: 'NOK'})
    df['Betrag okey'] = df['Total_rounded'].isin(valid_totals).map({True: 'OK', False: 'NOK'})
    
    return df

# applying the function
step2_df_only_TA = check_total_amounts(new_df[new_df["Faktor"] != 0].copy())


def check_nebenkosten(df):
    df = df.copy()
    
    total_rounded = df['Total'].round(2)
    
    df['Betrag okey'] = 'NOK'
    
    mask_leerfahrt = (
        df['InvoiceNr'].notna() & 
        df['InvoiceNr'].str.contains('Leerfahrt', case=False, na=False)
    )
    df.loc[mask_leerfahrt & total_rounded.isin([105.0, 52.5]), 'Betrag okey'] = 'OK'
    
    # Condition Nebenkosten = 'Terminzuschlag'
    mask_termin = df['Nebenkosten'] == 'Terminzuschlag'
    df.loc[mask_termin & total_rounded.isin([50.0, 100.0]), 'Betrag okey'] = 'OK'
    
    # condition: Nebenkosten = 'Seilwinde-Zuschlag'
    mask_seilwinde = df['Nebenkosten'] == 'Seilwinde-Zuschlag'
    df.loc[mask_seilwinde & (total_rounded == 50.0), 'Betrag okey'] = 'OK'
    
    # Condition other
    mask_other = (
        df['Nebenkosten'].notna() & 
        (df['Nebenkosten'] != 'Terminzuschlag') & 
        (df['Nebenkosten'] != 'Seilwinde-Zuschlag')
    )
    df.loc[mask_other & total_rounded.isin([12.70, 10.25]), 'Betrag okey'] = 'OK'
    
    return df

# applying
step2_df_only_nebenkocten = check_nebenkosten(new_df[new_df["Faktor"] == 0].copy())

# combining errors from both dataFrames Step2
errors_nebenkosten = step2_df_only_nebenkocten[step2_df_only_nebenkocten['Betrag okey'] == 'NOK'].copy()
errors_ta = step2_df_only_TA[step2_df_only_TA['Betrag okey'] == 'NOK'].copy() 
step2_df_errors = pd.concat([errors_nebenkosten, errors_ta], ignore_index=True)

# Step 3: counting nebencosten 
temp = step2_df_only_nebenkocten.copy()
temp['Total'] = temp['Total'].round(2)

# counting alle Nebencosten 
tv50_count = ((temp['Nebenkosten'] == 'Terminzuschlag') 
              & 
              (temp['Total'].isin([50.0]))).sum()

tv19_count = ((temp['Nebenkosten'].str.contains('termin', case=False, na=False) ) 
              & 
              (temp['Total'].isin([10.25, 19.0]))).sum()

tv100_count = ((temp['Nebenkosten'] == 'Terminzuschlag') & 
               (temp['Total'].isin([100.0, 55.0]))).sum()

leerfahrt_count = (temp['InvoiceNr'].notna() & 
                   temp['InvoiceNr'].str.contains('leerfahrt', case=False, na=False)).sum()

seilwinde_count = ((temp['Nebenkosten'] == 'Seilwinde-Zuschlag') & 
                   (temp['Total'] == 50.0)).sum()

eubername_count = (temp['Total'] == 12.70).sum()

# creating new DataFrame for Step 3 summary
summary_nebenkosten = pd.DataFrame({
    'Kategorie': ['TV50', 'TV19', 'TV100', 'Leerfahrt', 'Seilwinde', 'eÜbernahme'],
    'Anzahl': [tv50_count, tv19_count, tv100_count, leerfahrt_count, seilwinde_count, eubername_count]
})

# Step 4: checking alle Nebenkosten mit der Info aus der DB
# making copy of a dataFrame
step_4_nebenkosten = step2_df_only_nebenkocten[[
    "InvoiceNr",
    "Invoiceshort", 
    "Auftraggeber", 
    "VIN", 
    "Faktor", 
    "Nebenkosten", 
    "Total", 
    "Betrag okey"
]].copy()

# new columns (from the DB)
new_columns = [
    'Transport_WFP', 'WFP_Transportpreis (9010, 2010, 9020)', 'Transportcost', 'Telavis', 'Seilwinde', 
    'Terminzuschlag', 'EÜbernahme', 'Leerfahrt', 'Seilwindeintransport'
]
for col in new_columns:
    step_4_nebenkosten[col] = None

step_4_nebenkosten['Weiterverrechnet'] = ''

# Processing every row
for idx, row in step_4_nebenkosten.iterrows():
    if pd.isna(row["Invoiceshort"]):
        Invoiceshort = ""
    else:
        Invoiceshort = str(row["Invoiceshort"]).strip()
    
    # chosing ca3 or rrm
    if len(Invoiceshort) == 5:
        df_reference = db_rrm
    elif len(Invoiceshort) == 6:
        df_reference = db_ca3
    else:
        df_reference = None
    
    # looking for match in the DB table 
    if df_reference is not None:
        matching_rows = df_reference[df_reference["invoice"].astype(str).str.strip() == Invoiceshort]
        if len(matching_rows) > 0:
            ref_row = matching_rows.iloc[0]
        else:
            ref_row = None
    else:
        ref_row = None
    
    # New columns aus der DB filling
    if ref_row is not None:
        step_4_nebenkosten.at[idx, 'Transport_WFP'] = ref_row.get('Faktor', np.nan)
        step_4_nebenkosten.at[idx, 'WFP_Transportpreis (9010, 2010, 9020)'] = ref_row.get('Gallikerpreis', np.nan)
        step_4_nebenkosten.at[idx, 'Transportcost'] = ref_row.get('Transportcost', np.nan)
        step_4_nebenkosten.at[idx, 'Telavis'] = ref_row.get('Telavis', np.nan)
        step_4_nebenkosten.at[idx, 'Seilwinde'] = ref_row.get('Seilwinde', np.nan)
        step_4_nebenkosten.at[idx, 'Terminzuschlag'] = ref_row.get('Terminzuschlag', np.nan)
        step_4_nebenkosten.at[idx, 'EÜbernahme'] = ref_row.get('EÜbernahme', np.nan)
        step_4_nebenkosten.at[idx, 'Leerfahrt'] = ref_row.get('Leerfahrt', np.nan)
        step_4_nebenkosten.at[idx, 'Seilwindeintransport'] = ref_row.get('Seilwindeintransport', np.nan)
    
    # The logic for the column Weiterverrechnet
    total_rounded = round(row['Total'], 2) if pd.notna(row['Total']) else None
    nebkosten = row['Nebenkosten'] if pd.notna(row['Nebenkosten']) else ""
    invoice_nr = row['InvoiceNr'] if pd.notna(row['InvoiceNr']) else ""
    
    weiterverrechnet = 'Nein, bitte manuell prüfen'
    
    # Condition 1: Terminzuschlag (регистронезависимо)
    if pd.notna(nebkosten) and 'termin' in str(nebkosten).lower():
        termin_val = ref_row.get('Terminzuschlag') if ref_row is not None else None
        if pd.notna(total_rounded) and pd.notna(termin_val):
            weiterverrechnet = 'Ja'
    
    # Condition 2: Seilwinde-Zuschlag
    elif nebkosten == 'Seilwinde-Zuschlag':
        seilwinde_val = ref_row.get('Seilwinde') if ref_row is not None else None
        seil_trans_val = ref_row.get('Seilwindeintransport') if ref_row is not None else None
        if pd.notna(total_rounded) and (pd.notna(seilwinde_val) or pd.notna(seil_trans_val)):
            weiterverrechnet = 'Ja'
    
    # Condition 3: Leerfahrt 
    elif 'leerfahrt' in invoice_nr.lower():
        if pd.notna(total_rounded) and pd.notna(ref_row.get('Leerfahrt') if ref_row is not None else None):
            weiterverrechnet = 'Ja'
    
    # Condition 4: EÜbernahme / Seilwindeintransport (Total = 12.70)
    elif total_rounded is not None and abs(total_rounded - 12.70) < 0.01:
        euber = ref_row.get('EÜbernahme') if ref_row is not None else None
        if pd.notna(euber):
            weiterverrechnet = 'Ja'
    
    step_4_nebenkosten.at[idx, 'Weiterverrechnet'] = weiterverrechnet

# Statistik for checking
print("Statistik Weiterverrechnet:")
print(f"\nRows processed: {len(step_4_nebenkosten)}")

# STEP 5: Doppelte Werte
df_5 = df_5.loc[:, ~df_5.columns.str.startswith('Unnamed')]
df6 = df_5.loc[:, ~df_5.columns.str.startswith('Unnamed')]
if 'Fahrgestellnummer' in df_5.columns and 'Faktor' in df_5.columns:
    df_clean = df_5[df_5['Fahrgestellnummer'].notna() & df_5['Faktor'].notna()]
    duplicates_mask = df_clean.duplicated(subset=['Fahrgestellnummer', 'Faktor'], keep=False)
    
    if duplicates_mask.any():
        df_duplicates = df_clean[duplicates_mask].copy()
        print(f"Gefunden: {len(df_duplicates)} Doppeleinträge")
    else:
        df_duplicates = pd.DataFrame({'Meldung': ['Keine Doppeleinträge gefunden']})
        print("Keine Doppeleinträge gefunden")
else:
    df_duplicates = pd.DataFrame({
        'Meldung': ['Fehler: Spalten "Fahrgestellnummer" oder "Faktor" nicht vorhanden'],
        'Vorhandene_Spalten': [', '.join(df_5.columns.tolist())]
    })
    print("Fehler: Benötigte Spalten nicht gefunden")

if 'Meldung' in df_duplicates.columns:
    print("\nErgebnis:")
    print(df_duplicates.to_string(index=False))
else:
    display_cols = ['Fahrgestellnummer', 'Faktor']
    for col in ['Invoiceshort', 'InvoiceNr', 'Total', 'Auftraggeber']:
        if col in df_duplicates.columns:
            display_cols.append(col)
            break 
    
    print("\nGefundene Doppeleinträge:")
    print(df_duplicates[display_cols].head(20).to_string(index=False))

# STEP 6: Total überprüfen
total_row = df6[df6['Faktor'].astype(str).str.contains('Total', case=False, na=False)]

if len(total_row) > 0:
    row = total_row.iloc[0]
    
    invoice_value = row.get('Faktor', row.get('Faktor', 'N/A'))
    betrag_str = row.get('Betrag', '0')
    
    try:
        if isinstance(betrag_str, str):
            total_invoice_value = float(betrag_str.replace(',', '.'))
        else:
            total_invoice_value = float(betrag_str) 
    except (ValueError, TypeError):
        total_invoice_value = 0.0

    kalk_sum = new_df['Total'].sum()
    
    df_6 = pd.DataFrame({
        'Invoice': [invoice_value],
        'Total_invoice': [total_invoice_value],
        'Kalkulatorische_Betragssumme': [kalk_sum]
    })
    
    total_rounded = round(df_6['Total_invoice'].iloc[0], 2)
    kalk_rounded = round(df_6['Kalkulatorische_Betragssumme'].iloc[0], 2)
    
    df_6['Stimmt der Gesamtbetrag'] = 'Ja' if abs(total_rounded - kalk_rounded) < 0.01 else 'Nein'
    
    print(f"Total aus Rechnung: {total_rounded:.2f}")
    print(f"Kalkuliert: {kalk_rounded:.2f}")
    print(f"Vergleich: {df_6['Stimmt der Gesamtbetrag'].iloc[0]}")
else:
    df_6 = pd.DataFrame({
        'Invoice': ['N/A'],
        'Total_invoice': [0],
        'Kalkulatorische_Betragssumme': [new_df['Total'].sum()],
        'Stimmt der Gesamtbetrag': ['Fehler: "Total"-Zeile nicht gefunden']
    })
    print("Warnung: Keine Zeile mit 'Total' im Faktor gefunden")

# STEP 7: Komische Transporte
if 'Absender' in new_df.columns and 'Empfaenger' in new_df.columns:
    mask_absender = new_df['Absender'].notna() & new_df['Absender'].astype(str).str.contains('Gallik', case=False, na=False)
    mask_empfaenger = new_df['Empfaenger'].notna() & new_df['Empfaenger'].astype(str).str.contains('Gallik', case=False, na=False)
    mask = mask_absender & mask_empfaenger
    
    if mask.any():
        df_7 = new_df[mask].copy()
    else:
        df_7 = pd.DataFrame({'Meldung': ['Keine "komische" Transporte gefunden']})
else:
    df_7 = pd.DataFrame({
        'Meldung': ['Fehler: Spalten "Absender" oder "Empfaenger" nicht vorhanden'],
        'Vorhandene_Spalten': [', '.join(new_df.columns.tolist())]
    })

# STEP 8: Wieviele PW, SUV und LNF gab es
pw_count = len(new_df[new_df['Faktor'] == 1.0])
suv_count = len(new_df[new_df['Faktor'] == 1.5])
lnf_count = len(new_df[new_df['Faktor'].isin([2.0, 2.5])])
total_count = pw_count + suv_count + lnf_count

df_8 = pd.DataFrame({
    'PW': [pw_count],
    'SUV': [suv_count],
    'LNF': [lnf_count],
    'Total': [total_count]
})

step1_total = step1_df['CA3'].iloc[0] + step1_df['RRM'].iloc[0] + step1_df['Fehler'].iloc[0]
df_8['Check'] = 'OK' if df_8['Total'].iloc[0] == step1_total else 'NOK'

# opening excel file
with pd.ExcelWriter("file4.xlsx", engine="openpyxl") as writer:
    
    # writing Fehlerreport sheet
    # new_df.to_excel(writer, sheet_name="Fehlerreport", index=False)
    # worksheet = writer.sheets["Fehlerreport"]
    # for col_num in range(1, len(new_df.columns) + 1):
    #     worksheet.column_dimensions[get_column_letter(col_num)].width = 25

    # writing CA3 sheet
    # df_ca3.to_excel(writer, sheet_name="CA3", index=False)
    # worksheet = writer.sheets["CA3"]
    # for col_num in range(1, len(df_ca3.columns) + 1):
    #     worksheet.column_dimensions[get_column_letter(col_num)].width = 15

    # writing RRM sheet
    # df_rrm.to_excel(writer, sheet_name="RRM", index=False)
    # worksheet = writer.sheets["RRM"]
    # for col_num in range(1, len(df_rrm.columns) + 1):
    #     worksheet.column_dimensions[get_column_letter(col_num)].width = 15
    
    # writing Vergleich sheet
    # df_vergleich.to_excel(writer, sheet_name="Vergleich", index=False)
    # worksheet = writer.sheets["Vergleich"]
    # for col_num in range(1, len(df_vergleich.columns) + 1):
    #     worksheet.column_dimensions[get_column_letter(col_num)].width = 15
    
    # writing Step1 sheet (number CA3, RRM, Fehler)
    step1_df.to_excel(writer, sheet_name="Step_1 Anzahl TA", index=False)
    worksheet = writer.sheets["Step_1 Anzahl TA"]
    for col_num in range(1, len(step1_df.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 25

    # writing Step1_1 sheet (Fehler)
    step_1_1_df.to_excel(writer, sheet_name="Step_1_1 Fehler", index=False)
    worksheet = writer.sheets["Step_1_1 Fehler"]
    for col_num in range(1, len(step_1_1_df.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 25
    
    # writing Step1_2 sheet (CA3)
    step_1_2_df.to_excel(writer, sheet_name="Step_1_2 CA3", index=False)
    worksheet = writer.sheets["Step_1_2 CA3"]
    for col_num in range(1, len(step_1_2_df.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 25
    
    # writing Step1_3 sheet (CA3)
    step_1_3_df.to_excel(writer, sheet_name="Step_1_3 RRM", index=False)
    worksheet = writer.sheets["Step_1_3 RRM"]
    for col_num in range(1, len(step_1_3_df.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 25

    # writing Step1_4 sheet (CA3)
    df_zusatz.to_excel(writer, sheet_name="Step_1_4 Zusätzlich", index=False)
    worksheet = writer.sheets["Step_1_4 Zusätzlich"]
    for col_num in range(1, len(df_zusatz.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 25

    # writing Step2 alle Falschbeträge 
    step2_df_errors.to_excel(writer, sheet_name="Step_2_Falschbeträge", index=False)
    worksheet = writer.sheets["Step_2_Falschbeträge"]
    for col_num in range(1, len(step2_df_errors.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 25

    # writing Step2 sheet TA
    # step2_df_only_TA.to_excel(writer, sheet_name="Step_2_1_TA_Beträge", index=False)
    # worksheet = writer.sheets["Step_2_1_TA_Beträge"]
    # for col_num in range(1, len(step2_df_only_TA.columns) + 1):
    #     worksheet.column_dimensions[get_column_letter(col_num)].width = 25
    
    # writing Step2 sheet Nebenkosten
    # step2_df_only_nebenkocten.to_excel(writer, sheet_name="Step_2_2_Nebenkosten", index=False)
    # worksheet = writer.sheets["Step_2_2_Nebenkosten"]
    # for col_num in range(1, len(step2_df_only_nebenkocten.columns) + 1):
    #     worksheet.column_dimensions[get_column_letter(col_num)].width = 25
    
    # writing Step3 sheet Nebenkosten counting
    summary_nebenkosten.to_excel(writer, sheet_name="Step_3_Nebenkosten", index=False)
    worksheet = writer.sheets["Step_3_Nebenkosten"]
    for col_num in range(1, len(summary_nebenkosten.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 25
    
    # writing Step4 sheet Nebenkosten validieren mit WFP aus CA£ oder RRM
    step_4_nebenkosten.to_excel(writer, sheet_name="Step_4_Weiteverrechnet", index=False)
    worksheet = writer.sheets["Step_4_Weiteverrechnet"]
    for col_num in range(1, len(step_4_nebenkosten.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 25

    # writing Step5 sheet PW, SUV, LNF zählen
    df_8.to_excel(writer, sheet_name="Step_5_PW_SUV_LNF", index=False)
    worksheet = writer.sheets["Step_5_PW_SUV_LNF"]
    for col_num in range(1, len(df_8.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 25
   
    # writing Step6 sheet doppelte Werte ausfindig zu machen 
    df_duplicates.to_excel(writer, sheet_name="Step_6_Dublikate", index=False)
    worksheet = writer.sheets["Step_6_Dublikate"]
    for col_num in range(1, len(df_duplicates.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 25
    
    # writing Step7 sheet Gesamtbetrag überprüfen
    df_6.to_excel(writer, sheet_name="Step_7_Gesamtbetrag", index=False)
    worksheet = writer.sheets["Step_7_Gesamtbetrag"]
    for col_num in range(1, len(df_6.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 35
    
    # writing Step8 sheet Komische Transporte
    df_7.to_excel(writer, sheet_name="Step_8_Komische_Transporte", index=False)
    worksheet = writer.sheets["Step_8_Komische_Transporte"] 
    for col_num in range(1, len(df_7.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 25
    
    
