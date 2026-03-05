import os
import pandas as pd
from openpyxl.utils import get_column_letter
import json
import math
import httpx

# Input
filename = os.environ.get("INPUT_EXCEL_PATH")
if not filename:
    raise RuntimeError("INPUT_EXCEL_PATH is required for pdf5.py")

filepath = filename if os.path.isabs(filename) else os.path.join(os.getcwd(), filename)

try:
    df = pd.read_excel(filepath, skiprows=4)
    df['Betrag'] = df['Betrag'].astype(str).str.replace(',', '.').astype(float)
    df['Erledigt'] = pd.to_datetime(df['Erledigt'], format='%d.%m.%Y', errors='coerce')
    print("Datei geladen:", filepath)
except FileNotFoundError:
    raise RuntimeError(f"Datei nicht gefunden: {filepath}")

new_df = df[['Hersteller', 'Fahrgestellnummer', 'Fahrzeugtyp', 'Erledigt', 'Betrag']].copy()
new_df.rename(columns={
    'Fahrgestellnummer': 'VIN',
    'Fahrzeugtyp':       'Fahrzeug'
}, inplace=True)

new_df = new_df[new_df['Fahrzeug'].notna()]
new_df = new_df[new_df['Fahrzeug'].str.strip() != ""]

# Load DB
url = os.getenv("ca3_ausweise_schilder")

if not url:
    config_path = os.path.join(os.getcwd(), "config.json")
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        url = config.get("ca3_ausweise_schilder", "")

df_metabase = pd.read_json(url)
print("DB geladen:", len(df_metabase), "Einträge")

# Merge by VIN
merged_df = pd.merge(
    new_df,
    df_metabase,
    left_on="VIN",
    right_on="vin",
    how="left"
)

# Helpers
def is_empty(v):
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return False

# VIN_vergleich
merged_df["VIN_vergleich"] = merged_df["vin"].apply(
    lambda v: "NOK" if is_empty(v) else "OK"
)

# WFP vergleich (all 4)
WFP_COLS = ["wfp3300", "wfp3020", "wfp3018", "wfp3019"]

for wfp in WFP_COLS:
    merged_df[f"{wfp}_vergleich"] = merged_df.apply(
        lambda row, w=wfp: (
            "OK" if row["VIN_vergleich"] == "OK" and not is_empty(row.get(w))
            else "NOK"
        ),
        axis=1
    )

# Bemerkungen
def bemerkungen_logic(row):
    if row["VIN_vergleich"] == "NOK":
        return "Transportauftrag nicht gefunden. Bitte WFPs 2010, 9010, 9020, 3300, 3020, 3018, 3019 manuell prüfen"
    # Check if none of the 4 WFPs is activated
    all_nok = all(row[f"{w}_vergleich"] == "NOK" for w in WFP_COLS)
    if all_nok:
        return "nicht weiterverrechnet"
    return ""

merged_df["Bemerkungen"] = merged_df.apply(bemerkungen_logic, axis=1)

# Output: ALL rows
df_final = merged_df[[
    "Hersteller", "VIN", "Fahrzeug", "Erledigt", "Betrag",
    "Auftraggeber",
    "wfp3300", "wfp3020", "wfp3018", "wfp3019",
    "Bemerkungen"
]].copy()

BASE_DIR = os.environ.get("BASE_DIR")
if not BASE_DIR:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

final_file = os.path.join(BASE_DIR, "Fehlerreport.xlsx")

with pd.ExcelWriter(final_file, engine="openpyxl") as writer:
    df_final.to_excel(writer, index=False, sheet_name="Fehlerreport")
    ws = writer.sheets["Fehlerreport"]
    for i in range(1, len(df_final.columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.column_dimensions[get_column_letter(len(df_final.columns))].width = 50  # Bemerkungen

print(f"Fehlerreport gespeichert: {final_file} ({len(df_final)} Zeilen)")